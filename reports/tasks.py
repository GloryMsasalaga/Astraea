from celery import shared_task
from django.conf import settings
from django.utils import timezone
from django.core.files.base import ContentFile
from django.db.models import Sum, Count, Q
from io import BytesIO
import os
import logging
from datetime import datetime, timedelta
from decimal import Decimal

# PDF generation imports
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib import colors
from reportlab.graphics.shapes import Drawing
from reportlab.graphics.charts.piecharts import Pie
from reportlab.graphics.charts.linecharts import HorizontalLineChart
from reportlab.graphics.charts.barcharts import VerticalBarChart

# Excel generation imports
try:
    import openpyxl
    from openpyxl.chart import PieChart, BarChart, LineChart, Reference
    from openpyxl.styles import Font, Alignment, PatternFill
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

# HTML to PDF conversion
try:
    import weasyprint
    WEASYPRINT_AVAILABLE = True
except (ImportError, OSError):
    WEASYPRINT_AVAILABLE = False

from .models import (
    ReportTemplate, GeneratedReport, ReportSection, ReportChart
)
from documents.models import Document, ExtractedField
from reconciliation.models import ReconciliationSession, TransactionMatch, ReconciliationException
from dashboard.models import FinancialMetric, CashflowEntry, ExpenseCategory

logger = logging.getLogger(__name__)


@shared_task(bind=True)
def generate_report(self, report_id):
    """Generate a report based on template and parameters"""
    try:
        report = GeneratedReport.objects.get(id=report_id)
        report.status = 'processing'
        report.progress = 0
        report.save()
        
        template = report.template
        parameters = report.parameters
        
        # Update progress
        self.update_state(state='PROGRESS', meta={'current': 10, 'total': 100})
        report.progress = 10
        report.save()
        
        # Generate report content based on template type
        if template.template_type == 'financial_summary':
            content = generate_financial_summary_report(report, parameters)
        elif template.template_type == 'reconciliation_summary':
            content = generate_reconciliation_summary_report(report, parameters)
        elif template.template_type == 'document_analysis':
            content = generate_document_analysis_report(report, parameters)
        elif template.template_type == 'audit_trail':
            content = generate_audit_trail_report(report, parameters)
        elif template.template_type == 'expense_analysis':
            content = generate_expense_analysis_report(report, parameters)
        else:
            raise ValueError(f"Unknown template type: {template.template_type}")
        
        # Update progress
        self.update_state(state='PROGRESS', meta={'current': 50, 'total': 100})
        report.progress = 50
        report.save()
        
        # Generate file based on format
        if report.format == 'pdf':
            file_content = generate_pdf_report(content, template, parameters)
            file_extension = 'pdf'
            content_type = 'application/pdf'
        elif report.format == 'html':
            file_content = generate_html_report(content, template, parameters)
            file_extension = 'html'
            content_type = 'text/html'
        elif report.format == 'excel' and EXCEL_AVAILABLE:
            file_content = generate_excel_report(content, template, parameters)
            file_extension = 'xlsx'
            content_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        else:
            raise ValueError(f"Unsupported format: {report.format}")
        
        # Update progress
        self.update_state(state='PROGRESS', meta={'current': 80, 'total': 100})
        report.progress = 80
        report.save()
        
        # Save file
        timestamp = timezone.now().strftime('%Y%m%d_%H%M%S')
        filename = f"{report.title}_{timestamp}.{file_extension}"
        
        report.file_path.save(
            filename,
            ContentFile(file_content),
            save=False
        )
        report.file_size = len(file_content)
        report.status = 'completed'
        report.progress = 100
        report.completed_at = timezone.now()
        report.save()
        
        logger.info(f"Report {report_id} generated successfully")
        return {'status': 'completed', 'file_path': report.file_path.url}
        
    except Exception as e:
        logger.error(f"Error generating report {report_id}: {str(e)}")
        report.status = 'failed'
        report.error_message = str(e)
        report.save()
        raise


def generate_financial_summary_report(report, parameters):
    """Generate financial summary report content"""
    date_from = parameters.get('date_from')
    date_to = parameters.get('date_to')
    
    if date_from:
        date_from = datetime.strptime(date_from, '%Y-%m-%d').date()
    if date_to:
        date_to = datetime.strptime(date_to, '%Y-%m-%d').date()
    
    # Get financial metrics
    metrics_query = FinancialMetric.objects.filter(user=report.user)
    if date_from:
        metrics_query = metrics_query.filter(period_start__gte=date_from)
    if date_to:
        metrics_query = metrics_query.filter(period_end__lte=date_to)
    
    metrics = metrics_query.order_by('-period_start')
    
    # Calculate totals
    total_revenue = sum(m.revenue for m in metrics)
    total_expenses = sum(m.expenses for m in metrics)
    net_profit = total_revenue - total_expenses
    
    # Get cashflow data
    cashflow_query = CashflowEntry.objects.filter(user=report.user)
    if date_from:
        cashflow_query = cashflow_query.filter(date__gte=date_from)
    if date_to:
        cashflow_query = cashflow_query.filter(date__lte=date_to)
    
    cashflow_entries = cashflow_query.order_by('date')
    
    # Generate expense breakdown
    expense_breakdown = {}
    for entry in cashflow_entries.filter(entry_type='outflow'):
        category = entry.category or 'Other'
        expense_breakdown[category] = expense_breakdown.get(category, 0) + entry.amount
    
    content = {
        'title': report.title,
        'period': f"{date_from} to {date_to}" if date_from and date_to else "All time",
        'summary': {
            'total_revenue': total_revenue,
            'total_expenses': total_expenses,
            'net_profit': net_profit,
            'profit_margin': (net_profit / total_revenue * 100) if total_revenue > 0 else 0
        },
        'metrics': list(metrics.values()),
        'cashflow': list(cashflow_entries.values()),
        'expense_breakdown': expense_breakdown,
        'generated_at': timezone.now()
    }
    
    # Create report sections
    create_report_sections(report, content)
    
    return content


def generate_reconciliation_summary_report(report, parameters):
    """Generate reconciliation summary report content"""
    session_ids = parameters.get('session_ids', [])
    date_from = parameters.get('date_from')
    date_to = parameters.get('date_to')
    
    # Get reconciliation sessions
    sessions_query = ReconciliationSession.objects.filter(user=report.user)
    if session_ids:
        sessions_query = sessions_query.filter(id__in=session_ids)
    if date_from:
        date_from = datetime.strptime(date_from, '%Y-%m-%d').date()
        sessions_query = sessions_query.filter(created_at__date__gte=date_from)
    if date_to:
        date_to = datetime.strptime(date_to, '%Y-%m-%d').date()
        sessions_query = sessions_query.filter(created_at__date__lte=date_to)
    
    sessions = sessions_query.order_by('-created_at')
    
    # Calculate statistics
    total_sessions = sessions.count()
    completed_sessions = sessions.filter(status='completed').count()
    total_matches = 0
    total_exceptions = 0
    
    session_summaries = []
    for session in sessions:
        matches = TransactionMatch.objects.filter(
            ledger_record__session=session,
            is_confirmed=True
        ).count()
        exceptions = ReconciliationException.objects.filter(
            session=session
        ).count()
        
        total_matches += matches
        total_exceptions += exceptions
        
        session_summaries.append({
            'id': str(session.id),
            'name': session.name,
            'status': session.status,
            'matches': matches,
            'exceptions': exceptions,
            'created_at': session.created_at
        })
    
    content = {
        'title': report.title,
        'period': f"{date_from} to {date_to}" if date_from and date_to else "All time",
        'summary': {
            'total_sessions': total_sessions,
            'completed_sessions': completed_sessions,
            'completion_rate': (completed_sessions / total_sessions * 100) if total_sessions > 0 else 0,
            'total_matches': total_matches,
            'total_exceptions': total_exceptions,
            'match_rate': (total_matches / (total_matches + total_exceptions) * 100) if (total_matches + total_exceptions) > 0 else 0
        },
        'sessions': session_summaries,
        'generated_at': timezone.now()
    }
    
    # Create report sections
    create_report_sections(report, content)
    
    return content


def generate_document_analysis_report(report, parameters):
    """Generate document analysis report content"""
    date_from = parameters.get('date_from')
    date_to = parameters.get('date_to')
    document_types = parameters.get('document_types', [])
    
    # Get documents
    documents_query = Document.objects.filter(user=report.user)
    if date_from:
        date_from = datetime.strptime(date_from, '%Y-%m-%d').date()
        documents_query = documents_query.filter(uploaded_at__date__gte=date_from)
    if date_to:
        date_to = datetime.strptime(date_to, '%Y-%m-%d').date()
        documents_query = documents_query.filter(uploaded_at__date__lte=date_to)
    if document_types:
        documents_query = documents_query.filter(document_type__in=document_types)
    
    documents = documents_query.order_by('-uploaded_at')
    
    # Calculate statistics
    total_documents = documents.count()
    processed_documents = documents.filter(status='completed').count()
    failed_documents = documents.filter(status='failed').count()
    
    # Document type breakdown
    type_breakdown = {}
    for doc in documents:
        doc_type = doc.document_type or 'Unknown'
        type_breakdown[doc_type] = type_breakdown.get(doc_type, 0) + 1
    
    # Processing time analysis
    processing_times = []
    for doc in documents.filter(status='completed', completed_at__isnull=False):
        if doc.completed_at and doc.uploaded_at:
            processing_time = (doc.completed_at - doc.uploaded_at).total_seconds()
            processing_times.append(processing_time)
    
    avg_processing_time = sum(processing_times) / len(processing_times) if processing_times else 0
    
    content = {
        'title': report.title,
        'period': f"{date_from} to {date_to}" if date_from and date_to else "All time",
        'summary': {
            'total_documents': total_documents,
            'processed_documents': processed_documents,
            'failed_documents': failed_documents,
            'success_rate': (processed_documents / total_documents * 100) if total_documents > 0 else 0,
            'avg_processing_time': avg_processing_time
        },
        'type_breakdown': type_breakdown,
        'documents': list(documents.values()),
        'generated_at': timezone.now()
    }
    
    # Create report sections
    create_report_sections(report, content)
    
    return content


def generate_expense_analysis_report(report, parameters):
    """Generate expense analysis report content"""
    date_from = parameters.get('date_from')
    date_to = parameters.get('date_to')
    categories = parameters.get('categories', [])
    
    # Get expense data
    expenses_query = CashflowEntry.objects.filter(
        user=report.user,
        entry_type='outflow'
    )
    if date_from:
        date_from = datetime.strptime(date_from, '%Y-%m-%d').date()
        expenses_query = expenses_query.filter(date__gte=date_from)
    if date_to:
        date_to = datetime.strptime(date_to, '%Y-%m-%d').date()
        expenses_query = expenses_query.filter(date__lte=date_to)
    if categories:
        expenses_query = expenses_query.filter(category__in=categories)
    
    expenses = expenses_query.order_by('-date')
    
    # Calculate totals by category
    category_totals = {}
    monthly_totals = {}
    
    for expense in expenses:
        category = expense.category or 'Other'
        category_totals[category] = category_totals.get(category, 0) + expense.amount
        
        month_key = expense.date.strftime('%Y-%m')
        if month_key not in monthly_totals:
            monthly_totals[month_key] = {}
        monthly_totals[month_key][category] = monthly_totals[month_key].get(category, 0) + expense.amount
    
    total_expenses = sum(category_totals.values())
    
    content = {
        'title': report.title,
        'period': f"{date_from} to {date_to}" if date_from and date_to else "All time",
        'summary': {
            'total_expenses': total_expenses,
            'total_categories': len(category_totals),
            'avg_monthly_expense': total_expenses / len(monthly_totals) if monthly_totals else 0
        },
        'category_totals': category_totals,
        'monthly_totals': monthly_totals,
        'expenses': list(expenses.values()),
        'generated_at': timezone.now()
    }
    
    # Create report sections
    create_report_sections(report, content)
    
    return content


def generate_audit_trail_report(report, parameters):
    """Generate audit trail report content"""
    # This would include all user activities, changes, etc.
    # For now, we'll create a basic implementation
    content = {
        'title': report.title,
        'summary': {
            'total_activities': 0,
            'period': "All time"
        },
        'activities': [],
        'generated_at': timezone.now()
    }
    
    # Create report sections
    create_report_sections(report, content)
    
    return content


def create_report_sections(report, content):
    """Create report sections and charts from content"""
    
    # Create summary section
    ReportSection.objects.create(
        report=report,
        name="Executive Summary",
        content=content.get('summary', {}),
        section_type='summary',
        order=1
    )
    
    # Create charts based on content type
    if 'expense_breakdown' in content:
        chart_data = [
            {'label': category, 'value': float(amount)}
            for category, amount in content['expense_breakdown'].items()
        ]
        ReportChart.objects.create(
            report=report,
            title="Expense Breakdown",
            chart_type='pie',
            data=chart_data,
            order=1
        )
    
    if 'category_totals' in content:
        chart_data = [
            {'label': category, 'value': float(amount)}
            for category, amount in content['category_totals'].items()
        ]
        ReportChart.objects.create(
            report=report,
            title="Category Analysis",
            chart_type='bar',
            data=chart_data,
            order=2
        )


def generate_pdf_report(content, template, parameters):
    """Generate PDF report using ReportLab"""
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    styles = getSampleStyleSheet()
    story = []
    
    # Title
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        spaceAfter=30,
        textColor=colors.HexColor('#2c3e50')
    )
    story.append(Paragraph(content['title'], title_style))
    story.append(Spacer(1, 12))
    
    # Period
    if 'period' in content:
        period_style = ParagraphStyle(
            'Period',
            parent=styles['Normal'],
            fontSize=12,
            textColor=colors.HexColor('#7f8c8d')
        )
        story.append(Paragraph(f"Period: {content['period']}", period_style))
        story.append(Spacer(1, 20))
    
    # Summary section
    if 'summary' in content:
        story.append(Paragraph("Summary", styles['Heading2']))
        summary_data = []
        for key, value in content['summary'].items():
            if isinstance(value, (int, float)):
                if key.endswith('_rate') or key.endswith('_margin'):
                    formatted_value = f"{value:.2f}%"
                else:
                    formatted_value = f"${value:,.2f}" if 'amount' in key or 'revenue' in key or 'expense' in key else f"{value:,}"
            else:
                formatted_value = str(value)
            summary_data.append([key.replace('_', ' ').title(), formatted_value])
        
        summary_table = Table(summary_data, colWidths=[3*inch, 2*inch])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#34495e')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#ecf0f1')),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        story.append(summary_table)
        story.append(Spacer(1, 20))
    
    # Generated timestamp
    story.append(Spacer(1, 30))
    timestamp_style = ParagraphStyle(
        'Timestamp',
        parent=styles['Normal'],
        fontSize=10,
        textColor=colors.HexColor('#95a5a6'),
        alignment=2  # Right align
    )
    story.append(Paragraph(f"Generated on: {content['generated_at'].strftime('%Y-%m-%d %H:%M:%S')}", timestamp_style))
    
    doc.build(story)
    buffer.seek(0)
    return buffer.getvalue()


def generate_html_report(content, template, parameters):
    """Generate HTML report"""
    html_content = f"""
    <!DOCTYPE html>
    <html>
    <head>
        <title>{content['title']}</title>
        <style>
            body {{ font-family: Arial, sans-serif; margin: 40px; }}
            .header {{ border-bottom: 2px solid #2c3e50; padding-bottom: 20px; margin-bottom: 30px; }}
            .title {{ color: #2c3e50; font-size: 28px; margin: 0; }}
            .period {{ color: #7f8c8d; font-size: 14px; margin-top: 10px; }}
            .section {{ margin-bottom: 30px; }}
            .section h2 {{ color: #34495e; border-bottom: 1px solid #ecf0f1; padding-bottom: 10px; }}
            .summary-table {{ width: 100%; border-collapse: collapse; }}
            .summary-table th, .summary-table td {{ padding: 12px; text-align: left; border-bottom: 1px solid #ddd; }}
            .summary-table th {{ background-color: #34495e; color: white; }}
            .footer {{ margin-top: 50px; padding-top: 20px; border-top: 1px solid #ecf0f1; color: #95a5a6; font-size: 12px; }}
        </style>
    </head>
    <body>
        <div class="header">
            <h1 class="title">{content['title']}</h1>
            <div class="period">Period: {content.get('period', 'All time')}</div>
        </div>
        
        <div class="section">
            <h2>Summary</h2>
            <table class="summary-table">
                <thead>
                    <tr>
                        <th>Metric</th>
                        <th>Value</th>
                    </tr>
                </thead>
                <tbody>
    """
    
    # Add summary data
    if 'summary' in content:
        for key, value in content['summary'].items():
            if isinstance(value, (int, float)):
                if key.endswith('_rate') or key.endswith('_margin'):
                    formatted_value = f"{value:.2f}%"
                else:
                    formatted_value = f"${value:,.2f}" if 'amount' in key or 'revenue' in key or 'expense' in key else f"{value:,}"
            else:
                formatted_value = str(value)
            
            html_content += f"""
                    <tr>
                        <td>{key.replace('_', ' ').title()}</td>
                        <td>{formatted_value}</td>
                    </tr>
            """
    
    html_content += f"""
                </tbody>
            </table>
        </div>
        
        <div class="footer">
            Generated on: {content['generated_at'].strftime('%Y-%m-%d %H:%M:%S')}
        </div>
    </body>
    </html>
    """
    
    return html_content.encode('utf-8')


def generate_excel_report(content, template, parameters):
    """Generate Excel report using openpyxl"""
    if not EXCEL_AVAILABLE:
        raise ValueError("Excel generation not available - openpyxl not installed")
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Report Summary"
    
    # Title
    ws['A1'] = content['title']
    ws['A1'].font = Font(size=16, bold=True)
    ws['A1'].fill = PatternFill(start_color="2c3e50", end_color="2c3e50", fill_type="solid")
    
    # Period
    ws['A3'] = f"Period: {content.get('period', 'All time')}"
    ws['A3'].font = Font(size=12)
    
    # Summary section
    if 'summary' in content:
        ws['A5'] = "Summary"
        ws['A5'].font = Font(size=14, bold=True)
        
        row = 6
        ws['A6'] = "Metric"
        ws['B6'] = "Value"
        ws['A6'].font = Font(bold=True)
        ws['B6'].font = Font(bold=True)
        
        for key, value in content['summary'].items():
            row += 1
            ws[f'A{row}'] = key.replace('_', ' ').title()
            if isinstance(value, (int, float)):
                ws[f'B{row}'] = value
                if key.endswith('_rate') or key.endswith('_margin'):
                    ws[f'B{row}'].number_format = '0.00%'
                elif 'amount' in key or 'revenue' in key or 'expense' in key:
                    ws[f'B{row}'].number_format = '$#,##0.00'
            else:
                ws[f'B{row}'] = str(value)
    
    # Save to bytes
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()
